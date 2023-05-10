import openpyxl
import numpy as np
import datetime
import matplotlib.pyplot as plt
import tkinter as tk
from tkinter import simpledialog

#User Input
ROOT = tk.Tk()
ROOT.withdraw()
fname = simpledialog.askstring(title="File Location",
                                  prompt="Input file path:")
active_mass = simpledialog.askfloat(title="Active Mass",
                                  prompt="Input active mass:")
theo_speccap = simpledialog.askfloat(title="Theoretical Specific Capacity",
                                  prompt="Input theoretical specific capacity:")
lower_voltlim = simpledialog.askfloat(title="Voltage Range",
                                  prompt="Input lower voltage limit:")
upper_voltlim = simpledialog.askfloat(title="Voltage Range",
                                  prompt="Input upper voltage limit:")

figsize = (8,6)

class DataImport:
    def __init__(self, filename):
        self.filename = filename
    
    def load_sort(filename):
        
        def load(filename):
            # Loads data into python
            book = openpyxl.load_workbook(filename)
            sheet_obj = book.active
            max_col = sheet_obj.max_column
            col_headers = []
            for i in range(1, max_col + 1):
                cell_obj = sheet_obj.cell(row = 1, column = i)
                col_headers.append(cell_obj.value)
            data = {}
            for h in col_headers:
                data[h] = []
            for row in sheet_obj:
                for i, h in enumerate(col_headers):
                    data[h].append(row[i].value)
            for k, v in data.items():
                data[k] = np.array(v[1:])
            cycles = data['Cycle C']     
            data_types['TestTime'] = datetime.datetime
            data_types['StepTime'] = datetime.datetime
            return data
        
        def sort(data):
            # Sorts data via dictionaries in dictionaries in dictionaries...
            keys = list(data.keys())
            formatted_cycles = {}
            n_cycles = len(np.unique(data['Cycle C']))
            for i in range(n_cycles):
                formatted_cycles[i+1] = {'charge':{}, 'discharge':{}}
                for k in keys:
                    formatted_cycles[i+1]['charge'][k] = []
                    formatted_cycles[i+1]['discharge'][k] = []
            n_entries = len(data['Cycle C'])
            for i in range(n_entries):
                cyclenumber = data["Cycle C"][i]
                state = data['Md'][i]
                if state == 'C':
                    state = 'charge'
                elif state == 'D':
                    state = 'discharge'
                else:
                    continue
                for k in keys:
                    value = data[k][i]
                    formatted_cycles[cyclenumber][state][k].append(value)
            return formatted_cycles
        
        data_types = {'Rec': int,
                      'Cycle P': int,
                      'Cycle C': int,
                      'Step':int,
                      'TestTime': datetime.datetime,
                      'StepTime': datetime.datetime,
                      'Cap. [Ah]': float,
                      'Ener. [Wh]': float,
                      'Current [A]': float,
                      'Voltage [V]': float,
                      'Md': str,
                      'ES': int}
        
        data = load(filename)
        load_sort_data = sort(data)
        return load_sort_data

def Echem_Plots(filename):
    
    def cycling():
        data = DataImport.load_sort(filename)
        fig = plt.figure(figsize=figsize)
        for cyclenumber, cycledata in data.items():
            try:
                ccy = (cycledata['charge']['Cap. [Ah]'][-1])*1000/active_mass
                ddy = (cycledata['discharge']['Cap. [Ah]'][-1])*1000/active_mass
                plt.scatter(cyclenumber, ccy, s=80, facecolors='r', edgecolors='r')
                plt.scatter(cyclenumber, ddy, s=80, facecolors='none', edgecolors='r')
            except IndexError:
                continue
        plt.title('Specific Capacity vs. Cycle #')
        plt.xlabel('Cycle #')
        plt.ylabel('Specific Capacity (mAh/g)')
        plt.legend(["Charge", "Discharge"], loc ="upper right")
    
    def voltage_profiles():
        data = DataImport.load_sort(filename)
        n_cycles = len(data)
        agrad = np.linspace(0.3,0.8,n_cycles+1)
        
        def volt_vs_time():
            fig = plt.figure(figsize=figsize)
            for cyclenumber, cycledata in data.items():
                cx = cycledata['charge']['StepTime']
                cy = cycledata['charge']['Voltage [V]']
                dx = cycledata['discharge']['StepTime']
                dy = cycledata['discharge']['Voltage [V]']
                plt.plot(cx, cy, color='red', alpha=agrad[cyclenumber])
                plt.plot(dx, dy, color='blue', alpha=agrad[cyclenumber])
            plt.title('Voltage vs. Cycle Time')
            plt.xlabel('Time (hours)')
            plt.ylabel('Voltage (V)')
            plt.ylim(lower_voltlim, upper_voltlim)
            plt.margins(0)
            return fig
        
        def volt_vs_cap():
            fig = plt.figure(figsize=figsize)
            for cyclenumber, cycledata in data.items():
                v_cx = np.array(cycledata['charge']['Cap. [Ah]'])*1000/active_mass
                v_dx = np.array(cycledata['discharge']['Cap. [Ah]'])*1000/active_mass
                v_cy = cycledata['charge']['Voltage [V]']
                v_dy = cycledata['discharge']['Voltage [V]']
                plt.plot(v_cx, v_cy, color='red', alpha=agrad[cyclenumber])
                plt.plot(v_dx, v_dy, color='blue', alpha=agrad[cyclenumber])
            plt.title('Voltage vs. Specific Capacity')
            plt.xlabel('Specific Capacity (mAh/g)')
            plt.ylabel('Voltage (V)')
            plt.ylim(lower_voltlim, upper_voltlim)
            plt.margins(0)
            return fig
        
        def ee():
            fig = plt.figure(figsize=figsize)
            for cyclenumber, cycledata in data.items():
                v_cx = (np.array(cycledata['charge']['Cap. [Ah]'])*1000/active_mass)/theo_speccap
                v_dx = (np.array(cycledata['discharge']['Cap. [Ah]'])*1000/active_mass)/theo_speccap
                v_cy = cycledata['charge']['Voltage [V]']
                v_dy = cycledata['discharge']['Voltage [V]']
                plt.plot(v_cx, v_cy, color='red', alpha=agrad[cyclenumber])
                plt.plot(v_dx, v_dy, color='blue', alpha=agrad[cyclenumber])
            plt.title('Voltage vs. Electron Equivalence')
            plt.xlabel('Electron Equivalence')
            plt.ylabel('Voltage (V)')
            plt.xlim(0,1,0.2)
            plt.ylim(lower_voltlim, upper_voltlim)
            plt.margins(0)
            return fig
        
        volt_vs_time()
        volt_vs_cap()
        ee()
        
    cycling()
    voltage_profiles()

Echem_Plots(fname)
plt.show()
